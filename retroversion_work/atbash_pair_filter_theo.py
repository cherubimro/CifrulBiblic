#!/usr/bin/env python3
"""
Filter the v2 Atbash-pair scan xlsx: keep only pairs where AT LEAST ONE
side is a theologically-anchored Greek form (Jesus, Christ, Peter, Paul,
Father, Spirit, Lord, cross, blood, disciple, etc.) and print the top
pairs grouped by residue value.

This extracts interpretable findings without privileging a theological
residue dictionary — the residue can be anything.
"""
import sys
import openpyxl
import unicodedata
from collections import defaultdict
from pathlib import Path

WORK = Path('/home/bu/Documents/Biblia/retroversion_work')


def strip_accents(word):
    w = unicodedata.normalize('NFD', word.lower())
    return ''.join(c for c in w if unicodedata.category(c) != 'Mn')


# Greek theological anchor roots (lowercase, no accents). A pair is "anchored"
# if one of the two forms starts with one of these roots OR equals one of
# these stems exactly.
ANCHOR_ROOTS = [
    'ιησου',   # Ἰησοῦς/Ἰησοῦν/Ἰησοῦ
    'χριστ',   # Χριστός and variants
    'πετρ',    # Πέτρος/ν/υ
    'παυλ',    # Παῦλος
    'πατερ',   # πατέρα
    'πατηρ',   # πατήρ
    'πατρ',    # πατρός, πατρί
    'πνευμ',   # πνεύμα/τος
    'κυρι',    # κύριος
    'μαθητ',   # μαθητής/ήν
    'σταυρ',   # σταυρός/ωθείς
    'αιμ',     # αἷμα (but also αἰών? no, αἰων has no μ immediately)
    'λυτρ',    # λύτρον/ωτής
    'δουλ',    # δοῦλος (servant, as in Phil 2:7)
    'αναστ',   # ἀνάστασις
    'γραφ',    # γραφή (scripture)
    'αρχιερε', # ἀρχιερεύς (high priest)
    'ιερε',    # ἱερεύς (priest)
    'ναω',     # ναός (temple, gen. ναοῦ)
    'ναο',     # ναοῦ, ναόν, ναῷ
    'θυσι',    # θυσία (sacrifice)
    'πασχ',    # πάσχα
    'βαπτ',    # βαπτίζω/βάπτισμα
    'αμν',     # ἀμνός (lamb) - but also ἀμήν ? no, ἀμήν isn't rooted here
    'προφητ',  # προφήτης
    'αποστολ', # ἀπόστολος
    'εκκλησ',  # ἐκκλησία
    'βασιλει', # βασιλεία (kingdom) - NB: βασιλεύς too, but handled below
    'βασιλευ', # βασιλεύς
    'σωτηρ',   # σωτήρ
    'ηλι',     # Ἠλίας (Elijah)
    'μωυσ',    # Μωυσῆς
    'δαυιδ',   # Δαυίδ (wait this doesn't match lowercase accent strip — δαυίδ → δαυιδ)
    'ιωαν',    # Ἰωάννης
    'μαρια',   # Μαρία (wait won't work as prefix — Μαρία, Μαρίαμ)
    'ιουδ',    # Ἰούδας (also Judah)
    'λογ',     # λόγος — but very common
    'θεο',     # θεός — extremely common
    'αρνι',    # ἀρνίον (lamb, Revelation)
    'φως',     # φῶς (light)
    'ζω',      # ζωή (life)
    'αλεθ',    # ἀλήθεια (truth)  (typo: should be αληθ)
    'αληθ',    # ἀλήθεια
    'οδος',    # ὁδός (way)
    'ισχυ',    # ἰσχύς (strength)
    'χαρ',     # χάρις (grace)
    'ειρην',   # εἰρήνη (peace)
    'αγαπ',    # ἀγάπη
    'ελπ',     # ἐλπίς (hope)
    'πιστ',    # πίστις
    'μαρτυρ',  # μαρτυρία
]


def is_anchored(form):
    """Check if the form starts with or equals an anchor root (stripped)."""
    stripped = strip_accents(form)
    for root in ANCHOR_ROOTS:
        if stripped.startswith(root):
            return root
    return None


def main():
    path = WORK / 'atbash_pair_scan_v2.xlsx'
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    header = [c.value for c in ws[1]]

    # Columns: (see scan v2)
    # 0:A 1:A_lem 2:A_ct 3:A_iso 4:A_ref 5:B 6:B_lem 7:B_ct 8:B_iso 9:B_ref
    # 10:asum 11:csum 12:res 13:res_ct 14:resA 15:resB 16:iso_m 17:heb_m 18:score

    anchored = []
    for r in rows:
        A, B = r[0], r[5]
        rootA = is_anchored(A)
        rootB = is_anchored(B)
        if rootA or rootB:
            anchored.append(r + (rootA or '', rootB or ''))

    print(f'Total anchored pairs: {len(anchored)}', file=sys.stderr)

    # Group by residue
    by_residue = defaultdict(list)
    for r in anchored:
        by_residue[r[12]].append(r)

    # Rank residues by: (a) total pairs in anchored subset, (b) residue word
    # must be reasonably rare in NT overall (res_total <= 50 for signal)
    print('\n=== Residue values in anchored pairs (rank: total pairs, rarity) ===\n')

    # For each residue, print top 3 anchored pairs by A+B frequency
    sorted_res = sorted(by_residue.items(), key=lambda x: -len(x[1]))
    for res, pairs in sorted_res[:50]:
        # Skip very common residue values (noise)
        sample_res_total = pairs[0][13]
        if sample_res_total > 80:
            continue

        # Rank pairs: at least one side has count >= 5 (so at least one
        # "anchored" side is meaningful frequency), break ties by rarity of
        # the partner word
        def pair_key(p):
            ac, bc = p[2], p[7]
            # Prefer pairs where one side is frequent (anchor) and other is rare
            max_ct = max(ac, bc)
            min_ct = min(ac, bc)
            return (-max_ct, min_ct)
        pairs.sort(key=pair_key)

        # Print top 3 per residue
        top = pairs[:3]
        if not top:
            continue

        print(f'  residue={res:5}  (NT iso total freq × {sample_res_total:3}, '
              f'{len(pairs)} anchored pairs)')
        for p in top:
            iso_m = (p[16] or '').split(',')[0].strip() if p[16] else '-'
            heb_m = (p[17] or '').split(',')[0].strip() if p[17] else '-'
            print(f"    sum={p[10]:5}  {p[0]:18}×{p[2]:<4} ↔ "
                  f"{p[5]:18}×{p[7]:<4}  "
                  f"res_letters: {p[14]}/{p[15]}   → NT:{iso_m}")
        print()


if __name__ == '__main__':
    main()
