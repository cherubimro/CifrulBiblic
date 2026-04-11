#!/usr/bin/env python3
"""
Systematic NT numerical scan using retroversion.json + base_values.json.

For every Greek lemma in the NT:
  1. Compute isopsephy of each form
  2. Look up matches in base_values
  3. Compute Hebrew gematria of Delitzsch canonical stem
  4. Look up Hebrew matches
  5. Check factor 37 (Christ factor)
  6. Cross-language convergence: Greek form iso == Hebrew stem gem

Output:
  - nt_findings.xlsx (with all hits, ranked)
  - nt_findings_summary.md (human-readable top discoveries)

Filters out:
  - Known matches catalogued in gia/gematria_references.xlsx (marked "KNOWN")
  - Our own previous findings (marked "OUR_PRIOR")
"""
import json
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, '/home/bu/Documents/Biblia')

try:
    import openpyxl
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

WORK = Path('/home/bu/Documents/Biblia/retroversion_work')


def load_json(path):
    with open(path, encoding='utf-8') as f:
        return json.load(f)


def build_gia_lookup():
    """Load gia/gematria_references.xlsx into a {value: [entries]} lookup."""
    gia_path = Path('/home/bu/Documents/Biblia/gia/gematria_references.xlsx')
    if not gia_path.exists() or not HAS_XLSX:
        return {}
    wb = openpyxl.load_workbook(gia_path, read_only=True)
    ws = wb['References']
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    lookup = defaultdict(list)
    for r in rows:
        number = r[4]
        if number:
            try:
                val = int(number)
                lookup[val].append({
                    'word': r[7] or '',
                    'translit': r[8] or '',
                    'meaning': r[9] or '',
                    'by': r[11] or '',
                })
            except (TypeError, ValueError):
                pass
    return dict(lookup)


def has_factor_37(n):
    return n > 0 and n % 37 == 0


def factorize(n):
    """Return simple factorization string like '2 × 3 × 37'."""
    if n < 2:
        return str(n)
    factors = []
    d = 2
    x = n
    while d * d <= x:
        while x % d == 0:
            factors.append(d)
            x //= d
        d += 1
    if x > 1:
        factors.append(x)
    return ' × '.join(str(f) for f in factors)


def is_triangular(n):
    """Return k if n = T(k) = k*(k+1)/2, else None."""
    if n < 1:
        return None
    # k² + k - 2n = 0 → k = (-1 + sqrt(1+8n))/2
    import math
    k = (-1 + math.isqrt(1 + 8 * n)) / 2
    ki = int(k)
    if ki * (ki + 1) // 2 == n:
        return ki
    return None


def is_hexagonal(n):
    """Return k if n = H(k) = k*(2k-1), else None."""
    if n < 1:
        return None
    import math
    # 2k² - k - n = 0 → k = (1 + sqrt(1+8n))/4
    disc = 1 + 8 * n
    sd = math.isqrt(disc)
    if sd * sd != disc:
        return None
    k = (1 + sd) / 4
    ki = int(k)
    if ki * (2 * ki - 1) == n:
        return ki
    return None


def scan(retroversion, base_values, gia_lookup):
    findings = []

    for greek_lemma, entry in retroversion.items():
        ro = entry.get('ro', '')
        iso_lemma = entry.get('isopsephy_lemma')
        greek_forms = entry.get('greek_forms', [])
        canonical = entry.get('hebrew_canonical', {})
        candidates = entry.get('hebrew_candidates', [])

        # Source info for finding classification
        def classify(value):
            """Return ('KNOWN', source) if value is in gia_lookup OR base_values with known=True."""
            if value in gia_lookup:
                return 'KNOWN', gia_lookup[value][0].get('by', '')
            if str(value) in base_values:
                bv = base_values[str(value)]
                known = bv.get('known', False)
                if known is True:
                    return 'KNOWN', bv.get('source', '')
                elif isinstance(known, str):
                    return 'OUR_PRIOR', known
                else:
                    return 'NEW', ''
            return 'NONE', ''

        # === A. Greek form matches with base values ===
        for fidx, f in enumerate(greek_forms[:10]):  # top 10 forms
            v = f['iso']
            if str(v) in base_values:
                status, src = classify(v)
                findings.append({
                    'type': 'GREEK_FORM_HIT',
                    'value': v,
                    'greek_lemma': greek_lemma,
                    'greek_form': f['form'],
                    'form_count': f['count'],
                    'ro': ro,
                    'base_label': base_values[str(v)]['label'],
                    'status': status,
                    'source': src,
                    'factor_37': has_factor_37(v),
                    'factorization': factorize(v),
                    'hebrew_stem': canonical.get('stem', ''),
                    'hebrew_gem': canonical.get('gematria', 0),
                    'cross_match': (canonical.get('gematria') == v) if canonical else False,
                })

        # === B. Greek lemma factor-37 hits ===
        if has_factor_37(iso_lemma or 0) and iso_lemma and iso_lemma > 37:
            factor = iso_lemma // 37
            # Only include multiples up to 100×37 to avoid noise
            if factor <= 100:
                status, _ = classify(iso_lemma)
                findings.append({
                    'type': 'FACTOR_37_LEMMA',
                    'value': iso_lemma,
                    'greek_lemma': greek_lemma,
                    'greek_form': greek_lemma,
                    'form_count': sum(f['count'] for f in greek_forms),
                    'ro': ro,
                    'base_label': f'{factor} × 37',
                    'status': status,
                    'source': '',
                    'factor_37': True,
                    'factorization': factorize(iso_lemma),
                    'hebrew_stem': canonical.get('stem', ''),
                    'hebrew_gem': canonical.get('gematria', 0),
                    'cross_match': False,
                })

        # === C. Cross-language convergence: Greek form iso == Hebrew stem gem ===
        if canonical:
            heb_gem = canonical.get('gematria', 0)
            for fidx, f in enumerate(greek_forms[:10]):
                if f['iso'] == heb_gem and heb_gem > 0:
                    status, _ = classify(heb_gem)
                    findings.append({
                        'type': 'CROSS_LANG_MATCH',
                        'value': heb_gem,
                        'greek_lemma': greek_lemma,
                        'greek_form': f['form'],
                        'form_count': f['count'],
                        'ro': ro,
                        'base_label': f'{f["form"]} Greek = {canonical["stem"]} Hebrew',
                        'status': status,
                        'source': '',
                        'factor_37': has_factor_37(heb_gem),
                        'factorization': factorize(heb_gem),
                        'hebrew_stem': canonical.get('stem', ''),
                        'hebrew_gem': heb_gem,
                        'cross_match': True,
                    })

        # === D. Hebrew canonical hit on base value ===
        if canonical:
            hgem = canonical.get('gematria', 0)
            if str(hgem) in base_values:
                status, src = classify(hgem)
                findings.append({
                    'type': 'HEBREW_CANONICAL_HIT',
                    'value': hgem,
                    'greek_lemma': greek_lemma,
                    'greek_form': canonical.get('stem', ''),
                    'form_count': canonical.get('form_most_common', '') and 1 or 0,
                    'ro': ro,
                    'base_label': base_values[str(hgem)]['label'],
                    'status': status,
                    'source': src,
                    'factor_37': has_factor_37(hgem),
                    'factorization': factorize(hgem),
                    'hebrew_stem': canonical.get('stem', ''),
                    'hebrew_gem': hgem,
                    'cross_match': False,
                })

    return findings


def score_and_rank(findings):
    """Assign a score to each finding; higher = more interesting."""
    for f in findings:
        s = 0
        if f['type'] == 'CROSS_LANG_MATCH':
            s += 30
        if f['factor_37']:
            s += 10
        if f['status'] == 'NEW':
            s += 20
        elif f['status'] == 'OUR_PRIOR':
            s += 5
        elif f['status'] == 'KNOWN':
            s -= 5
        # Higher frequency = more important (caps at +10)
        s += min(f['form_count'], 100) // 10
        f['score'] = s
    findings.sort(key=lambda x: (-x['score'], x['type'], -x['value']))


def write_xlsx(findings, out_path):
    if not HAS_XLSX:
        print('openpyxl not available, skipping xlsx output')
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'NT Findings'
    headers = ['Score', 'Type', 'Status', 'Value', 'Factorization', 'Factor37',
               'Greek Lemma', 'Greek Form', 'Form Count', 'RO',
               'Hebrew Stem', 'Hebrew Gem', 'CrossMatch', 'Base Label', 'Source']
    ws.append(headers)
    for f in findings:
        ws.append([
            f['score'],
            f['type'],
            f['status'],
            f['value'],
            f['factorization'],
            'Y' if f['factor_37'] else '',
            f['greek_lemma'],
            f['greek_form'],
            f['form_count'],
            f['ro'],
            f['hebrew_stem'],
            f['hebrew_gem'],
            'Y' if f['cross_match'] else '',
            f['base_label'],
            f['source'],
        ])
    # Freeze header + auto width
    ws.freeze_panes = 'A2'
    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    wb.save(out_path)


def main():
    print('Loading data...')
    retroversion = load_json(WORK / 'retroversion.json')
    base_values = load_json(WORK / 'base_values.json')
    base_values = {k: v for k, v in base_values.items() if not k.startswith('_')}
    print(f'  Retroversion: {len(retroversion)} lemmas')
    print(f'  Base values: {len(base_values)}')

    print('Building gia lookup...')
    gia_lookup = build_gia_lookup()
    print(f'  gia references: {sum(len(v) for v in gia_lookup.values())} entries on {len(gia_lookup)} unique values')

    print('Scanning...')
    findings = scan(retroversion, base_values, gia_lookup)
    print(f'  Total raw findings: {len(findings)}')

    score_and_rank(findings)

    # Dedupe by (type, value, greek_lemma, greek_form)
    seen = set()
    uniq = []
    for f in findings:
        key = (f['type'], f['value'], f['greek_lemma'], f['greek_form'])
        if key not in seen:
            seen.add(key)
            uniq.append(f)
    print(f'  Deduped: {len(uniq)}')

    # Write xlsx
    out_xlsx = WORK / 'nt_findings.xlsx'
    write_xlsx(uniq, out_xlsx)
    print(f'\nWrote {out_xlsx}')

    # Stats by type and status
    from collections import Counter
    by_type = Counter(f['type'] for f in uniq)
    by_status = Counter(f['status'] for f in uniq)
    print(f'\nBy type: {dict(by_type)}')
    print(f'By status: {dict(by_status)}')

    # Top 30 findings
    print(f'\n=== Top 30 findings by score ===')
    for f in uniq[:30]:
        print(f"  [{f['score']:>3}] {f['type']:20} {f['status']:10} "
              f"{f['value']:>5} {f['greek_form']:15} "
              f"{f['hebrew_stem']:15} ({f['ro'][:25]}) — {f['base_label'][:50]}")


if __name__ == '__main__':
    main()
