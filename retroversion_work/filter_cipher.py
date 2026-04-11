#!/usr/bin/env python3
"""
Apply filters A (strict) and B (broad) to the raw cipher scan output,
plus focused analysis of key theological values.
"""
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

sys.path.insert(0, '/home/bu/Documents/Biblia')

try:
    import openpyxl
except ImportError:
    print('openpyxl required', file=sys.stderr)
    sys.exit(1)

WORK = Path('/home/bu/Documents/Biblia/retroversion_work')

# Key theological values to analyze separately
KEY_VALUES = {
    103: "Δανιηλ (Daniel) LXX / our Barabbas factor",
    148: "פסח (Pesach, Passover)",
    153: "John 21:11 fish / T(17) / H(9) / בני האלהים",
    206: "דבר (Word) / בר אבא (son of father)",
    207: "אור (light)",
    214: "רוח (ruach, spirit)",
    222: "888-666 / our Thomas 6×37",
    248: "אברהם (Abraham)",
    259: "βασιλεία (kingdom) / 7×37",
    276: "Acts 27:37 / T(23) / H(12)",
    358: "משיח (Messiah) / נחש (serpent)",
    385: "שכינה (Shekinah) / σινδόνα",
    391: "ישועה (salvation)",
    444: "κορβανᾶς / מקדש / 12×37",
    613: "γέγραφα / Torah 613 mitzvot",
    616: "Rev 13:18 variant",
    666: "Beast / Solomon",
    888: "Ἰησοῦς",
    911: "ראשית / χάρις",
    913: "בראשית",
    1209: "Atbash sum Ἰησοῦς/Πέτρος",
    1480: "Χριστός = 40×37",
}


def read_raw_cipher():
    wb = openpyxl.load_workbook(WORK / 'nt_ot_cipher.xlsx', read_only=True)
    ws = wb['NT↔OT Cipher']
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    cols = ['score', 'nt_form', 'iso', 'factor37', 'greek_lemma', 'ro',
            'nt_count', 'first_ref',
            'retroversion_stem', 'ot_hebrew_stem', 'ot_hebrew_strongs',
            'ot_occurrences', 'ot_first_verse']
    data = []
    for r in rows:
        d = dict(zip(cols, r))
        data.append(d)
    return data


def apply_filter_a(matches):
    """STRICT: iso >= 100, nt_count >= 3, ot_occ >= 20,
    and (iso in KEY_VALUES OR divisible by 37, 148, 74)."""
    out = []
    for m in matches:
        iso = m['iso'] or 0
        nt = m['nt_count'] or 0
        ot = m['ot_occurrences'] or 0
        if iso < 100 or nt < 3 or ot < 20:
            continue
        hit_reason = []
        if iso in KEY_VALUES:
            hit_reason.append(f"KEY:{iso}")
        for d in (37, 74, 148):
            if iso > d and iso % d == 0:
                hit_reason.append(f"÷{d}")
        if not hit_reason:
            continue
        m['filter_reason'] = ';'.join(hit_reason)
        out.append(m)
    return out


def apply_filter_b(matches):
    """BROAD: iso >= 50, nt_count >= 2, ot_occ >= 10."""
    out = []
    for m in matches:
        iso = m['iso'] or 0
        nt = m['nt_count'] or 0
        ot = m['ot_occurrences'] or 0
        if iso < 50 or nt < 2 or ot < 10:
            continue
        out.append(m)
    return out


def write_xlsx(matches, path, title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    headers = ['Score', 'NT Form', 'Iso', 'Factor37', 'Greek Lemma', 'RO',
               'NT #', 'First NT ref', 'Retroversion', 'OT Hebrew word',
               'Strongs H', 'OT #', 'OT first verse']
    if matches and 'filter_reason' in matches[0]:
        headers.append('Filter')
    ws.append(headers)
    for m in matches:
        row = [
            m['score'], m['nt_form'], m['iso'], m['factor37'],
            m['greek_lemma'], m['ro'], m['nt_count'], m['first_ref'],
            m['retroversion_stem'], m['ot_hebrew_stem'], m['ot_hebrew_strongs'],
            m['ot_occurrences'], m['ot_first_verse'],
        ]
        if 'filter_reason' in m:
            row.append(m['filter_reason'])
        ws.append(row)
    ws.freeze_panes = 'A2'
    for col in ws.columns:
        try:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
        except (ValueError, AttributeError):
            pass
    wb.save(path)


def analyze_key_values(matches):
    """For each key value, show all NT forms + all OT words at that value."""
    print('\n' + '=' * 80)
    print('KEY VALUE ANALYSIS')
    print('=' * 80)

    # Group matches by iso value
    by_iso = defaultdict(list)
    for m in matches:
        by_iso[m['iso'] or 0].append(m)

    for val, label in KEY_VALUES.items():
        if val not in by_iso:
            print(f'\n{val:>5} ({label}): NO MATCHES')
            continue
        ms = by_iso[val]
        # Unique NT forms and OT words
        nt_forms = {}
        ot_words = {}
        for m in ms:
            if m['nt_form'] not in nt_forms or (m['nt_count'] or 0) > (nt_forms[m['nt_form']]['nt_count'] or 0):
                nt_forms[m['nt_form']] = m
            if m['ot_hebrew_stem'] not in ot_words:
                ot_words[m['ot_hebrew_stem']] = m

        # Sort NT forms by count
        top_nt = sorted(nt_forms.values(), key=lambda x: -(x['nt_count'] or 0))[:10]
        # Sort OT words by occurrence
        top_ot = sorted(ot_words.values(), key=lambda x: -(x['ot_occurrences'] or 0))[:10]

        print(f'\n{val:>5} ({label})')
        print(f'  → NT forms at this value (top 10 by frequency):')
        for m in top_nt:
            print(f"      {m['nt_form']:20} (×{m['nt_count']:>3}, {m['greek_lemma']:12} = {(m['ro'] or '')[:25]:25}) @ {m['first_ref']}")
        print(f'  → OT Hebrew words at this value (top 10 by frequency):')
        for m in top_ot:
            print(f"      {m['ot_hebrew_stem']:15} (H{m['ot_hebrew_strongs']:<6} ×{m['ot_occurrences']:>4}) @ {m['ot_first_verse']}")


def main():
    print('Loading raw cipher matches...', file=sys.stderr)
    matches = read_raw_cipher()
    print(f'  Total raw: {len(matches)}', file=sys.stderr)

    fa = apply_filter_a(matches)
    fb = apply_filter_b(matches)
    print(f'  Filter A (strict, key/37/74/148 only): {len(fa)}', file=sys.stderr)
    print(f'  Filter B (broad): {len(fb)}', file=sys.stderr)

    write_xlsx(fa, WORK / 'nt_ot_cipher_strict.xlsx', 'Strict')
    write_xlsx(fb, WORK / 'nt_ot_cipher_broad.xlsx', 'Broad')
    print(f'Wrote nt_ot_cipher_strict.xlsx ({len(fa)} matches)', file=sys.stderr)
    print(f'Wrote nt_ot_cipher_broad.xlsx ({len(fb)} matches)', file=sys.stderr)

    # Key value focused analysis
    analyze_key_values(matches)

    # Top 30 from strict filter
    print('\n' + '=' * 80)
    print('TOP 30 — STRICT FILTER')
    print('=' * 80)
    for m in fa[:30]:
        print(f"  [{m['score']:>3}] {m['filter_reason']:15} {m['iso']:>5} | "
              f"{m['nt_form']:15} ({m['greek_lemma']:10} = {(m['ro'] or '')[:18]:18}) "
              f"↔ {m['ot_hebrew_stem']:12} "
              f"(H{m['ot_hebrew_strongs']}, ×{m['ot_occurrences']})")


if __name__ == '__main__':
    main()
