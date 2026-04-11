#!/usr/bin/env python3
"""
Re-rank cipher matches by statistical rarity.

For each NT↔OT match at value V:
  P(NT at V) = count_NT_at_V / total_NT_words
  P(OT at V) = count_OT_at_V / total_OT_words
  rarity(V) = 1 / (P(NT) × P(OT))

Higher rarity = match is statistically less likely by chance alone.
A match at V=666 is much rarer than V=385 because Hebrew has few words at 666.

The rarity score is multiplied by a "narrative weight" (semantic context,
NT form frequency, OT word frequency) to rank matches.

Output: nt_ot_cipher_rarity.xlsx — re-ranked top findings.
"""
import json
import math
import sys
from collections import Counter, defaultdict
from pathlib import Path

sys.path.insert(0, '/home/bu/Documents/Biblia')

try:
    import openpyxl
except ImportError:
    print('openpyxl required', file=sys.stderr)
    sys.exit(1)

from biblegematria.biblegematria import load_sblgnt, isopsephy

WORK = Path('/home/bu/Documents/Biblia/retroversion_work')


def load_json(path):
    with open(path, encoding='utf-8') as f:
        return json.load(f)


def compute_nt_value_counts():
    """Count NT word tokens per isopsephy value."""
    sblgnt = load_sblgnt()
    counts = Counter()
    for w in sblgnt:
        form = w['word'].strip('.,;·:()[]')
        v = isopsephy(form)
        if v > 0:
            counts[v] += 1
    return counts, sum(counts.values())


def compute_ot_value_counts():
    """Count OT word tokens per gematria_stem value."""
    morph = load_json(WORK / 'morph_index.json')
    counts = Counter()
    for vid, words in morph.items():
        for w in words:
            g = w['gematria_stem']
            if g > 0:
                counts[g] += 1
    return counts, sum(counts.values())


def read_raw_cipher():
    wb = openpyxl.load_workbook(WORK / 'nt_ot_cipher.xlsx', read_only=True)
    ws = wb['NT↔OT Cipher']
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    cols = ['score', 'nt_form', 'iso', 'factor37', 'greek_lemma', 'ro',
            'nt_count', 'first_ref',
            'retroversion_stem', 'ot_hebrew_stem', 'ot_hebrew_strongs',
            'ot_occurrences', 'ot_first_verse']
    return [dict(zip(cols, r)) for r in rows]


def main():
    print('Counting NT isopsephy values...', file=sys.stderr)
    nt_counts, nt_total = compute_nt_value_counts()
    print(f'  {len(nt_counts)} unique NT values, {nt_total} word tokens', file=sys.stderr)

    print('Counting OT gematria values...', file=sys.stderr)
    ot_counts, ot_total = compute_ot_value_counts()
    print(f'  {len(ot_counts)} unique OT values, {ot_total} word tokens', file=sys.stderr)

    # Compute rarity per value:
    # rarity(V) = -log10(P(NT at V) * P(OT at V))
    # This is "bits of improbability"
    rarity = {}
    for v in set(nt_counts.keys()) & set(ot_counts.keys()):
        p_nt = nt_counts[v] / nt_total
        p_ot = ot_counts[v] / ot_total
        r = -math.log10(p_nt * p_ot)
        rarity[v] = r

    print(f'\nRarity score range: {min(rarity.values()):.1f} – {max(rarity.values()):.1f}', file=sys.stderr)

    # Show rarity for key values
    key_vals = [103, 111, 148, 153, 206, 207, 214, 222, 248, 259, 276, 358, 385,
                391, 444, 613, 616, 666, 800, 848, 888, 911, 1209, 1480]
    print('\nRarity score for key values:', file=sys.stderr)
    for v in key_vals:
        nt_c = nt_counts.get(v, 0)
        ot_c = ot_counts.get(v, 0)
        if nt_c > 0 and ot_c > 0:
            r = rarity[v]
            print(f"  {v:>5}: NT×{nt_c:>4} × OT×{ot_c:>4} → rarity = {r:.1f} bits", file=sys.stderr)
        else:
            print(f"  {v:>5}: NT×{nt_c:>4} × OT×{ot_c:>4} → NO MATCH POSSIBLE", file=sys.stderr)

    # Load raw cipher matches
    print('\nLoading raw cipher matches...', file=sys.stderr)
    matches = read_raw_cipher()
    print(f'  {len(matches)} raw matches', file=sys.stderr)

    # Re-score
    print('Re-scoring by rarity...', file=sys.stderr)
    for m in matches:
        v = m['iso'] or 0
        r = rarity.get(v, 0)
        nt_freq = m['nt_count'] or 1
        ot_freq = m['ot_occurrences'] or 1
        # Narrative weight: favor forms that appear multiple times in NT
        # and OT words that are common enough to be "recognized"
        # (not hapax)
        narrative = math.log10(1 + nt_freq) + math.log10(1 + min(ot_freq, 1000))
        m['rarity'] = round(r, 2)
        m['narrative'] = round(narrative, 2)
        m['new_score'] = round(r * 10 + narrative * 5, 1)

    matches.sort(key=lambda x: -x['new_score'])

    # Write xlsx
    out_xlsx = WORK / 'nt_ot_cipher_rarity.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Rarity-ranked'
    headers = ['NewScore', 'Rarity(bits)', 'Narrative', 'NT Form', 'Iso',
               'Greek Lemma', 'RO', 'NT #', 'First NT ref',
               'OT Hebrew', 'Strongs H', 'OT #', 'OT first verse']
    ws.append(headers)
    for m in matches[:3000]:  # top 3000
        ws.append([
            m['new_score'], m['rarity'], m['narrative'],
            m['nt_form'], m['iso'],
            m['greek_lemma'], m['ro'], m['nt_count'], m['first_ref'],
            m['ot_hebrew_stem'], m['ot_hebrew_strongs'],
            m['ot_occurrences'], m['ot_first_verse'],
        ])
    ws.freeze_panes = 'A2'
    for col in ws.columns:
        try:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
        except (ValueError, AttributeError):
            pass
    wb.save(out_xlsx)
    print(f'\nWrote {out_xlsx} (top 3000 by rarity)', file=sys.stderr)

    # Show top 50
    print('\n=== TOP 50 BY STATISTICAL RARITY ===')
    seen_forms = set()
    count = 0
    for m in matches:
        # Deduplicate by (form, iso) — same form could match multiple OT words
        key = (m['nt_form'], m['iso'])
        if key in seen_forms:
            continue
        seen_forms.add(key)
        count += 1
        if count > 50:
            break
        print(f"  [{m['new_score']:>6.1f}] r={m['rarity']:>4.1f} | "
              f"{m['iso']:>5} | "
              f"{m['nt_form']:18} ×{m['nt_count']:<3} "
              f"({m['greek_lemma']:10} = {(m['ro'] or '')[:18]:18}) "
              f"↔ {m['ot_hebrew_stem']:10} "
              f"(H{m['ot_hebrew_strongs']}, ×{m['ot_occurrences']}) "
              f"@ {m['first_ref']}")


if __name__ == '__main__':
    main()
